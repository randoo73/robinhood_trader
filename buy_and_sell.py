import openpyxl
from os.path import exists

# Get list of stock names
    # Use the buy_records directory

stockList = [];

########################################################

def populateStockList():

    stockList.clear()
    wb = openpyxl.load_workbook('stocks.xlsx') # Loads the Spreadsheet

    # stockList = wb.sheetnames  # Loads the Sheet of the

    sL = wb.sheetnames
    for i in sL:
        stockList.append(i)

    wb.close()
    
    stockList.remove('Totals')

    return sL

########################################################

def find_average(i):
      
    # Check to see if the list of buys file exists
    my_path = exists("buy_records/" + i + ".txt")
    
    # print(i)
    
    #print(my_path)

    if my_path:
    
        # Get the average of the previous buys.
        # Make sure the average is less than the current Price before selling.
        f = open("buy_records/" + i + ".txt", "r")
        total1 = 0
        total2 = 0
        x2 = 0
        y2 = 0
        yAverage = 0
        count = 0
        l1 = []
        # Loop through each line of the text file
        for line in f:
            # Make sure there is something on the line
            if len(line) > 1:
                # Get the total number of lines to calculate the average
                count += 1
                # Get rid of special characters and white space
                l1 = line.split(",")
                # Get the Trade Share Dollar Amount
                x1 = float(l1[0])
                #print(x1)
                x2 = x2 + x1
                # Get the Trade Share Quantity
                y1 = float(l1[1])
                y2 = y2 + y1
                # Get the accumulated Trade Dollar Amount
                total1 = float(round((x1 * y1) + total1, 2))
                
        #print(x2)
        #print("Counter: ", count)
        # Get the average price
        averagePriceFloat = float(round(x2/count, 2))
        # Get the accumulated Trade Dollar Amount
        total1 = float(round(((x1 * y1) + total1), 2))
        # Calculate hte profit price
        profitPrice = round((averagePriceFloat * 1.05), 2)
        
        sharesOwned = round(y2, 4)

        yAverage = round(total1/count, 2)
        f.close()
        
        print("Stock: ",  i)               
        print(f"Shares Owned: {sharesOwned}")
        print(f"Average Price: {averagePriceFloat}")
        print(f"Price To Beat: {profitPrice}")
        print()
        return averagePriceFloat

########################################################


def stocks_owned(i):
      
    # Check to see if the list of buys file exists
    my_path = exists("buy_records/" + i + ".txt")

    if my_path:
    
        # Get the average of the previous buys.
        # Make sure the average is less than the current Price before selling.
        f = open("buy_records/" + i + ".txt", "r")
        total1 = 0
        total2 = 0
        x2 = 0
        y2 = 0
        yAverage = 0
        count = 0
        l1 = []
        # Loop through each line of the text file
        for line in f:
            # Make sure there is something on the line
            if len(line) > 1:
                # Get the total number of lines to calculate the average
                count += 1
                # Get rid of special characters and white space
                l1 = line.split(",")

                # Get the Trade Share Quantity
                y1 = float(l1[1])
                y2 = y2 + y1
                
                sharesOwned = round(y2, 4)

                f.close()

                return sharesOwned

########################################################


# Acquire and iterate through the list and print the averages.

# Calculate the profit point.

# populateStockList()

# print(stockList)
# print()

# for i in stockList:
    # find_average(i)