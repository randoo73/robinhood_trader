#! /usr/bin/env python3

# multipage_robinhood_multiple_stock 
#
# Need to prove the executions in the orderConfirmation() function works better.
#
# Need to fix the roundedSharesToBuy function.
#     This needs is currently hard coded at $20.
#     It needs to be flexible to allow for increasing the dollar amount purchased by 10 every ????
#     2000 dollars of balance?
# 
# Need to fix the currentPriceFloat in the recordSharesandPrice() function.  Need to create a way to record
# the original purchase price and take an average from that and the current price.
# Currently it is ((old price + new price) divided by 2) * 1.05
#
# Re-Create Database to include the following tables:
#    Open Orders
#    Closed Orders
# The Open Orders should be added to in the Buy and Buy More functions and deleted in the Sell function
#    It should include the following columns:
#        Transaction Number,
#        Company Name and Symbol,
#        Shares Purchased,
#        Shares Owned,
#        Purchase Price,
#        Transaction Cost,
#        Transaction Date
# The Closed Orders table should be added to in the Sell function.
#    It should include the original input from the Open Orders table as well as:
#        Closed Transaction Date
#        Sale Price
# The findMinimumTradeAmount() function needs to be re-written.
#        It is defaulted to $20
#
# File is stocks.xlsx
import os
import openpyxl    # For working with spreadsheets
import datetime
import time
import shutil    # To copy and move files
# import robin_stocks as rs
import csv    # For working with csv files
import robin_stocks.robinhood as rs
import sqlite3
from counter_new_day import transactionCounter
import backup_and_record
import buy_and_sell


# Checks to see if a file exists
from os.path import exists

import authenticate as a    # Login function

#########################################################################

# Create an empty list
stockList = [];
currentPriceList = [];
oldPriceList = [];
quantityOwnedList = [];
buySellList = [];
recordedList = [];
shouldBeRecordedList = [];
tradedList = [];
buyMoreModifiedPriceList = [];

# Variables
cellNumber = 0
filled = 'filled'
cancelled = 'cancelled'
unconfirmed = 'unconfirmed'
tC = 0     # Total Available Cash
minimumTradeAmount = 20

transactionCost = 0
lastTransactionCumulativeQuantity = 0

stockIndex = 0
currentPriceIndex = 0
oldPriceIndex = 0
quantityOwnedIndex = 0
buySellIndex = 0
buyMoreModified = 0
databaseCounter = 0


saturday = 'Saturday'
sunday = 'Sunday'

########################################################

def populateStockList():

    stockList.clear()
    wb = openpyxl.load_workbook('stocks.xlsx') # Loads the Spreadsheet

    # stockList = wb.sheetnames  # Loads the Sheet of the

    sL = wb.sheetnames
    for i in sL:
        stockList.append(i)

    wb.close()
    
    stockList.remove('Totals')

    return sL
########################################################

def populateCurrentPriceList():
    currentPriceList.clear() # Clears the list from the last run.

    for item in stockList:
        # Get current price
        # Retrieve share price from Robinood -- this only comes back as strings in a list
        stockPrice = rs.stocks.get_latest_price(item)
    
    # Pull the price out of the list
        stockStr = stockPrice[0]
    
    # Convert the string to a float
        stockFloat = float(stockStr)
        currentStkPrice = round(stockFloat, 2)
            # Append to a list
        currentPriceList.append(currentStkPrice)
########################################################

def populateOldPriceList():
    # Clears the list from the last run.
    oldPriceList.clear() 

    wb = openpyxl.load_workbook("stocks.xlsx")

    for item in stockList:
        # Loads the Sheet of the WorkBook
        ws = wb[item]

        # Loads the Cell of the worksheet
        cC = ws.cell(row = ws.max_row, column = 3)
        
        if(cC == type(None)):
            print(f"Missing Values: {item}")
            oldPriceList.append(currentPriceList[stockIndex])
        
        else:
            oldPriceList.append(cC.value)

    wb.close()

    return cC

########################################################

def populatebuyMoreModifiedPriceList():
    # Clears the list from the last run.
    buyMoreModifiedPriceList.clear() 

    wb = openpyxl.load_workbook("stocks.xlsx")

    for item in stockList:
        # Loads the Sheet of the WorkBook
        ws = wb[item]

        # Loads the Cell of the worksheet
        cD = ws.cell(row = ws.max_row, column = 3)
        
        if(cD == type(None)):
            print(f"Missing Values: {item}")
            buyMoreModifiedPriceList.append(currentPriceList[stockIndex])
        
        else:
            buyMoreModifiedPriceList.append(cD.value)

    wb.close()

    return cD
########################################################

def populateQuantityOwnedList():
    # Clears the list from the last run.
    quantityOwnedList.clear() 

    wb = openpyxl.load_workbook("stocks.xlsx", data_only = True)

    for item in stockList:
        # Loads the Sheet of the WorkBook
        ws = wb[item]

        # Loads the Cell of the worksheet
        cE = ws.cell(row = ws.max_row, column = 6)
        cEv = cE.value
        
        # Incase nothing was recorded in this field
        if(cEv == type(None)):
            print(f"Missing Stock Data For: {item}")
            quantityOwnedList.append(0)
        else:
            cEvI = float(cEv)
            cEvIA = abs(cEvI)
            quantityOwnedList.append(cEvIA)

    wb.close()

    return cE
########################################################

def populateBuySellList():
    # Clears the list from the last run.
    buySellList.clear() 

    wb = openpyxl.load_workbook("stocks.xlsx", data_only = True)

    for item in stockList:
        # Loads the Sheet of the WorkBook
        ws = wb[item]

        # Loads the Cell of the worksheet
        cB = ws.cell(row = ws.max_row, column = 2)
        cBv = cB.value

        buySellList.append(cBv)

    wb.close()

    return cBv
########################################################

def getStockQuantity():
    
    quantityOwned = quantityOwnedList[stockIndex]
    quantityFloat = float(quantityOwned)
        
    return quantityFloat
########################################################

def tradeableCash():

    # Check available cash balance.
    my_profile = rs.profiles.load_account_profile()

    # format: print(dictionary[key])
    availableCash = my_profile["cash_available_for_withdrawal"]

    # Convert to float from string and find 1%
    availableCash1 = float(availableCash)

    aC1 = round(availableCash1, 2)

    return aC1
########################################################

def findMinimumTradeAmount():
    
    # Check to be sure the Available Cash is above $20
    if(minimumTradeAmount >= 20):
        
        # Make sure this minimum can go up as value grows
        # Make sure there is padding to trade with
        
        oldMinimum = 0
        
        newMinimumTradeAmount1 = int(tC)
        
        newMinimumTradeAmount2 = newMinimumTradeAmount1 / 2
        
        newMinimumTradeAmount3 = int(newMinimumTradeAmount2)
        
        with open('minimumTradeAmount.csv', newline='') as f:
            reader = csv.reader(f)
            for row in reader:
                print()

            oldMinimum = int(row[0])


        if(newMinimumTradeAmount2 > oldMinimum):
            minimumTradeAmount1 = newMinimumTradeAmount3
        else:
            minimumTradeAmount1 = oldMinimum

        with open('minimumTradeAmount.csv', 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(str(newMinimumTradeAmount3))

        return minimumTradeAmount1
    else:
        return minimumTradeAmount
########################################################

def stkPrice():
    
    # Retrieve share price from Robinood -- this only comes back as strings in a list
    stockPrice = rs.stocks.get_latest_price(stockList[stockIndex])
    
    # Pull the price out of the list
    stockStr = stockPrice[0]
    
    # Convert the string to a float
    stockFloat = float(stockStr)
    roundedStkPrice = round(stockFloat, 2)
    
    return roundedStkPrice
#########################################################

def roundedSharesToBuy():

    stkPrice1 = stkPrice()
    
    # tC is 1% of available cash
    #if (tC < 20):
        # Determine 1% or $20.00
    #    sharesToBuy = 20 / stkPrice1
    #else:
    #    sharesToBuy = tC / stkPrice1
    
    sharesToBuy = 20 / stkPrice1
    
    return round(sharesToBuy, 4)
########################################################

def orderConfirmation():
    
    # Checks to see if the order was filled

    getOrdersList = rs.orders.find_stock_orders(symbol = stockList[stockIndex], cancel= None)

    getOrders = getOrdersList[0]
    lastTransactionAveragePrice = getOrders['average_price']
    lastTransactionCumulativeQuantity = getOrders['cumulative_quantity']
    lastTransactionState1 = getOrders['state']
    lastTransactionState = lastTransactionState1.strip()
    lastTransactionTime = getOrders['last_transaction_at']
    transactionCost = getOrders['average_price']

    #print("Get Orders Length is: ", len(getOrders))
    print()
    print(f"Get Orders of orderConfirmation(): {getOrders}")
    print()
    
    x = lastTransactionTime
    y = x[:10]
    datemask = "%Y-%m-%d"
    z = datetime.datetime.strptime(y, datemask).date()

    print(f"Last Transaction State is: {lastTransactionState}")

    if(z != todayNumerical):
        print("Is not Today")
        print(f"DateStamp: {z}")
        print(f"Today: {today}")
        print()
        print("No Return Yet")
        time.sleep(5)
        
    if(z == todayNumerical):
        if(lastTransactionState is not cancelled):
            if(lastTransactionState is not filled):
                print(f"Order State is: {lastTransactionState}")
                print(f"DateStamp: {z}")
                print()
        else:
            filename = todayNumerical + '_' + stockList[stockIndex] + '_errorlog.txt'
            with open(filename, 'w') as f:
                f.write(getOrdersList)

        print(f"Order State is: {lastTransactionState}")
        print(f"DateStamp: {z}")
        print()

    return lastTransactionState

########################################################

def recordSharesandPrice(stockTradeAmount):
    
    # Keep a list of traded stocks to keep a count of trades this run
    
    tradedList.append(stockList[stockIndex])
    
    # Record Transaction data to the Spreadsheet
    
    # Retrieve order info
    currentPriceFloat = 0
    getOrdersList = rs.orders.find_stock_orders(symbol = stockList[stockIndex], cancel= None)

    getOrders = getOrdersList[0]
    executions = getOrders['executions']
    executions1 = executions[0]
    
    quantityOwned = quantityOwnedList[stockIndex]
    quantityOwnedFloat = float(quantityOwned)
    
    oldPrice = oldPriceList[stockIndex]
    oldPriceFloat = float(oldPrice)
    
    # Get Share Quantity
    sharesBoughtStr = stockTradeAmount
    sharesBoughtFloat = float(sharesBoughtStr)
    roundedSharesBought = round(sharesBoughtFloat, 4)
    ownedShareFloat = (quantityOwned + roundedSharesBought)

    # Get Share Price
    currentPrice = currentPriceList[stockIndex]
    currentPriceFloat = float(currentPrice)
        
    # Calculate the Buy More Price to Record
    if(buyMore == True):
        print(f"BuyMore currentPriceFloat: {currentPriceFloat}")
        print()    

    # Loads the Spreadsheet
    wb = openpyxl.load_workbook("stocks.xlsx")
       # Loads the Sheet of the WorkBook
    ws = wb[stockList[stockIndex]]
    time.sleep(0.1)
    
# Need to Record:
    # Column A = Todays date
    cA = ws.cell(row = ws.max_row + 1, column = 1)
    strToday = str(todayNumerical)
    cA.value = strToday
    print(f"Recorded Date is: {strToday}")
    time.sleep(0.1)
    shouldBeRecordedList.append(strToday)

    # Column B = Buy or Sell
    buyOrSellString = buySellList[stockIndex]
    print(f"Buy or Sell String: {buyOrSellString}")
    cB = ws.cell(row = ws.max_row, column = 2)
    if((buyOrSellString == 'Buy') and (currentPriceFloat < oldPriceFloat)):  # Determine if it is a BuyMore Entry
        cB.value = 'Buy'
        print(f"Recorded Condition is: {cB}")
        shouldBeRecordedList.append(cB.value)
    if(buyOrSellString == 'Sell'):
        cB.value = 'Buy'
        print(f"Recorded Condition is: {cB}")
        shouldBeRecordedList.append(cB.value)
    if((buyOrSellString == 'Buy') and (currentPriceFloat > oldPriceFloat)):
        cB.value = 'Sell'
        print(f"Recorded Condition is: {cB}")
        shouldBeRecordedList.append(cB.value)
    time.sleep(0.1)
    

    # Column C = Transaction Price
    # If it is a BuyMore:
    if((buyOrSellString == 'Buy') and (currentPriceFloat < oldPriceFloat)):  # Determine if it is a BuyMore Entry
        
        recordPrice1 = ws.cell(row = ws.max_row, column = 3)  # Last Transaction Price is found in Column 3
        recordPrice1.value = currentPriceFloat
        
        recordPrice = ws.cell(row = ws.max_row, column = 4)# Buy More Modified Price is found in Column 4
        recordPrice.value = currentPriceFloat
        
        print(f"Rounded Stock Price is: {currentPriceFloat}")
        print(f"Buy More Modified Stock Price is: {currentPriceFloat}")
        time.sleep(0.1)
        shouldBeRecordedList.append(currentPriceFloat)
        shouldBeRecordedList.append(currentPriceFloat)
    
    else:        
        recordPrice = ws.cell(row = ws.max_row, column = 3)  # Last Transaction Price is found in Column 3
        recordPrice.value = currentPriceFloat
        print(f"Rounded Stock Price is: {currentPriceFloat}")
        time.sleep(0.1)
        shouldBeRecordedList.append(currentPriceFloat)

    # Column E and F = Transaction Quantity
 
    # Shares Bought -- Negative if Buy, Positive if Sell
    recordShares = ws.cell(row = ws.max_row, column = 5)  # Transaction Share Quantities are found in Column 4
    recordShares1 = ws.cell(row = ws.max_row, column = 6) # This is what populateQuantityOwned() reads from
    time.sleep(0.1)
    
    # Current Transaction is a Buy
    if(buyOrSellString =='Sell'):
        recordShares.value = (roundedSharesBought * -1)
        recordShares1.value = (roundedSharesBought * -1)
        shouldBeRecordedList.append(roundedSharesBought * -1)
        shouldBeRecordedList.append(roundedSharesBought * -1)
    
    # Current transaction is a Sell
    if((buyOrSellString == 'Buy') and (currentPriceFloat > oldPriceFloat)):
        recordShares.value = roundedSharesBought
        recordShares1.value = 0
        shouldBeRecordedList.append(roundedSharesBought)
        shouldBeRecordedList.append(0)
        
    
    # Current transaction is a Buy More
    if((buyOrSellString == 'Buy') and (currentPriceFloat < oldPriceFloat)):
        recordShares.value = (roundedSharesBought * -1)
        recordShares1.value = ((roundedSharesBought + quantityOwnedFloat) * -1)
        shouldBeRecordedList.append(roundedSharesBought * -1)
        shouldBeRecordedList.append((roundedSharesBought + quantityOwnedFloat) * -1)
    time.sleep(0.1)

    # Column H = Transaction Price * Transaction Quantity
    recordTotal = ws.cell(row = ws.max_row, column = 8) # Transaction Cost is found in Column 7
    time.sleep(0.1)
    
    transactionCost = round((currentPriceFloat * roundedSharesBought), 2)
    print()
    print(f"Transaction Cost: {transactionCost}")
    print()
    transactionCostFloat = float(transactionCost)
    
    if(buyOrSellString == 'Sell'):  # Determine if THis was a Buy Transaction
        recordTotal.value = (transactionCostFloat * -1)
        shouldBeRecordedList.append(transactionCostFloat * -1)
        backup_and_record.database_buy(stockList[stockIndex], currentPriceFloat, roundedSharesBought, ownedShareFloat, transactionCostFloat)
        
    if((buyOrSellString == 'Buy') and (currentPriceFloat < oldPriceFloat)):  # Determine if a Buy More
        recordTotal.value = (transactionCostFloat * -1)
        shouldBeRecordedList.append(transactionCostFloat * -1)
        backup_and_record.database_buy(stockList[stockIndex], currentPriceFloat, roundedSharesBought, ownedShareFloat, transactionCostFloat)
        
    if((buyOrSellString == 'Buy') and (currentPriceFloat > oldPriceFloat)):  # Determine if THis was a Sell Transaction
        recordTotal.value = transactionCostFloat
        shouldBeRecordedList.append(transactionCostFloat)
        backup_and_record.database_sell(stockList[stockIndex], currentPriceFloat, roundedSharesBought, transactionCost)
    
    time.sleep(0.1)

    print(f"Recorded {roundedSharesBought}", f"shares at location {recordShares}")
    print(f"Recorded {currentPriceFloat}", f"dollars at location {recordPrice}")
    print()
    print(f"Should Be Recorded List: {shouldBeRecordedList}")
    print()
    

    
    # Save the Changes to Disk
    wb.save('stocks.xlsx')
    time.sleep(0.1)
    wb.close()
    time.sleep(0.1)
    
    shouldBeRecordedList.clear()
    
    backup_and_record.backup()
    
    #if(buyOrSellString == 'Buy'):
    #    backup_and_record.database_buy(stockList[stockIndex], currentPriceFloat, roundedSharesBought, ownedShareFloat, transactionCostFloat)
    #else:
    #    backup_and_record.database_sell(stockList[stockIndex], currentPriceFloat, roundedSharesBought, transactionCost)
#######################################################

def tenSecondNap():
    i = 10
    while i > 0:
        print(i)
        time.sleep(1)
        i -= 1
########################################################

def oneHourNap():
    i = 60
    while i > 0:
        if(i % 5 == 0):
            print(i)
        time.sleep(60)
        i -= 1
########################################################

def sixHourNap():
    i = 6
    while i > 0:
        print(f"Hours Left: {i}")
        oneHourNap()
        i -= 1
########################################################

def buyStock():
    
    #print("Buy Stock Function")
   
    stockTradeAmount = roundedSharesToBuy()

    if (buySellList[stockIndex] == 'Sell'):
        print("Time to Buy")
        print(f"Purchase Quantity: {stockTradeAmount}", f"of {stockList[stockIndex]}")
        print()
        
        # Make purchase
        purchase = rs.orders.order_buy_fractional_by_quantity(stockList[stockIndex], stockTradeAmount, timeInForce = 'gfd', extendedHours = False)
        
        # print(f"Purchase ID: {purchase}") # Prove what the return looks like
            
        tenSecondNap()    # Give the system time to process the order
        
        orderConf = orderConfirmation()

        if(orderConf == filled):
            recordSharesandPrice(stockTradeAmount)
            # Record purchase into a txt file.  This will make the sell
            # calculation easier with Buy More's.
            currentPrice = currentPriceList[stockIndex]
            f = open("buy_records/" + stockList[stockIndex] + ".txt", "w")
            f.write(str(currentPrice) + "," + str(stockTradeAmount) + "\n")
            f.close()
        
        
        if(orderConf == cancelled):
            print("Order Cancelled -- Try Again On Next Run")
            print()
        if((orderConf != filled) and (orderConf != cancelled)):
            print("Unknown Error - Figure out how to get Error Data")
            print("Possibly Get Link from List and have Python OPen it")
    print()

########################################################

def sellStock():
    
    stockTradeAmount = buy_and_sell.stocks_owned(stockList[stockIndex])
        
    # Make the Sale
    sale = rs.orders.order_sell_fractional_by_quantity(stockList[stockIndex], stockTradeAmount, timeInForce = 'gfd', extendedHours = False)
    print(f"Sale ID: {sale}") # Prove what the return looks like

    tenSecondNap()    # Give the system time to process the order

    orderConf = orderConfirmation()

    #This will show the order has been filled
    if(orderConf == filled):
        # Record quantity purchased and price in Spreadsheet:
        recordSharesandPrice(stockTradeAmount)

    if(orderConf == cancelled):
        print("Order Cancelled -- Try Again On Next Run")
        print()
    if((orderConf != filled) and (orderConf != cancelled)):
        print("Unknown Error - Figure out how to get Error Data")
        print("Possibly Get Link from List and have Python Open it")
        print()

########################################################

def buyMoreStock():
    print("Buy More Function")
    
    oldPrice = oldPriceList[stockIndex]    
    oldPriceFloat = float(oldPrice)
    
    buyMorePrice = (oldPriceFloat * 0.95)
    
    stockTradeAmount = roundedSharesToBuy()
    
    currentPrice = currentPriceList[stockIndex]
    currentPriceFloat = float(currentPrice)
    
    if (currentPriceFloat <= buyMorePrice):
        print(f"Shares Owned: {getStockQuantity()}")
        print("Time To Buy More")
        print(f"Purchase Quantity: {stockTradeAmount}", f"of {stockList[stockIndex]}")
        print()
        
        # Purchase AQdditional Stock
        purchaseMore = rs.orders.order_buy_fractional_by_quantity(stockList[stockIndex], stockTradeAmount, timeInForce = 'gfd', extendedHours = False)
        print(f"Purchase More ID: {purchaseMore}") # Prove what the return looks like
        
        tenSecondNap()    # Give the system time to process the order

        orderConf = orderConfirmation()

        #This will show the order has been filled
        if(orderConf == filled):            
            # Record quantity purchased and price in Spreadsheet:
            recordSharesandPrice(stockTradeAmount)
            # Record purchase into a txt file.  This will make the sell
            # calculation easier with Buy More's.
            currentPrice = currentPriceList[stockIndex]
            f = open("buy_records/" + stockList[stockIndex] + ".txt", "a")
            f.write("\n" + str(currentPrice) + "," + str(stockTradeAmount) + "\n")
            f.close()
        if(orderConf == cancelled):
            print("Order Cancelled -- Try Again On Next Run")
            print()
        if((orderConf != filled) and (orderConf != cancelled)):
            print("Unknown Error - Figure out how to get Error Data")
            print("Possibly Get Link from List and have Python Open it")
    print()

    #wb.close()
########################################################

getDateTime = datetime.datetime.now()
today = getDateTime.strftime(" %b %d, %Y")
timeNow = getDateTime.strftime("%X")
getDay = getDateTime.strftime("%A")
print()
print(f"Today is: {getDay} {today}")
print(f"The Time is: {timeNow}")
print()


while True:
    
    starttime = 0
    getDateTime = datetime.datetime.now()
    starttime = getDateTime
    today = getDateTime.strftime(" %b %d, %Y")
    todayNumerical = getDateTime.strftime("%Y-%m-%d")
    getDay = getDateTime.strftime("%A")
    getHour = getDateTime.strftime("%H")
    timeNow = getDateTime.strftime("%X")
    getMinute = getDateTime.strftime("%M")
    
    if getHour == '09':
        pass
    #########################################################################
    
    if((getDay != saturday) and (getDay != sunday)):
        
        if((getHour >= '07') and (getHour <= '12') and (getMinute == '23')):
            print(f"Today is: {getDay} {today}")
            print(f"The Time is: {timeNow}")
            print()
            print("Beginning Run...")
            print()
            
            # Log in Verification
            a.log_in()
        
############################################################################
            
            stockList.clear() # Clears the list from the last run.
            currentPriceList.clear()
            oldPriceList.clear()
            quantityOwnedList.clear()
            recordedList.clear()
            shouldBeRecordedList.clear()
            tradedList.clear()
            buySellList.clear()
            buyMoreModifiedPriceList.clear()
            
            print()
            tC = tradeableCash()
            minimumTradeAmount = 20 # findMinimumTradeAmount()
            minimumTradeAmount1 = 100 * minimumTradeAmount
            print()
            print(f"** Will only Trade if available Cash is more than ${minimumTradeAmount1}")
            print()
            print("** All purchases are a minimum of $10.00.")
            print()

            print(f"StockList: {stockList}")
            print(f"Current Price List: {currentPriceList}")
            print(f"Old Price List: {oldPriceList}")
            print(f"Quantity Owned List: {quantityOwnedList}")
            print()
            populateStockList()
            populateCurrentPriceList()
            populateOldPriceList()
            populateQuantityOwnedList()
            populateBuySellList()
            tC = tradeableCash()
            minimumTradeAmount = 20 # findMinimumTradeAmount()
            populatebuyMoreModifiedPriceList()
            print()
            print(f"StockList: {stockList}")
            print()
            print(f"Current Price List: {currentPriceList}")
            print()
            print(f"Old Price List: {oldPriceList}")
            print()
            print(f"Buy More Modified Price List: {buyMoreModifiedPriceList}")
            print()
            print(f"Quantity Owned List: {quantityOwnedList}")
            print()
            print(f"Buys and Sells: {buySellList}")
            print()
            print(f"Currently Trading {len(stockList)} stocks.")
            print()

            buyMore = False
#########################################################################

            for item in stockList:
                recordedList.clear()
                shouldBeRecordedList.clear()
                
                stockIndex = stockList.index(item)
                
                # Get the quantity owned - the amount purchased by this program
                quantityOwned = buy_and_sell.stocks_owned(item)
                quantityFloat = float(quantityOwned)
                
                # Get the last purchase price
                oldPrice = oldPriceList[stockIndex]
                oldPriceFloat = round(float(oldPrice), 2)
                
                # Get the average price paid since the last sell
                averagePrice = buy_and_sell.find_average(item)
                averagePriceFloat = round(float(averagePrice), 2)
                
                # Get the stocks current price
                currentPrice = currentPriceList[stockIndex]
                currentPriceFloat = float(currentPrice)
                
                # Find the estimated Profits
                profit = ((averagePriceFloat * 1.05) * quantityFloat) - (averagePriceFloat * quantityFloat)
                profitRounded = round(profit, 2)

                tC = tradeableCash()
                minimumTradeAmount = 20 # findMinimumTradeAmount()
                print(f"Available Cash is: ${tC}")
                print(f"The Minimum Tradeable Balance is: ${minimumTradeAmount}")
                print()
                
                print(f"{item}'s Current Price is: ${currentPriceFloat}")
                print(f"{item}'s Old Price was: ${oldPriceFloat}")
                print(f"Quantity Currently Owned of {item} is: {quantityFloat}")
                print(f"Will sell at: ${round(((averagePriceFloat) * 1.05), 2)} for a profit of: ${profitRounded}")
                print(f"Or buy more at: ${round(((oldPriceFloat) * 0.90), 2)}")
                print()

                if(buySellList[stockIndex] == 'Sell'):
                    if(tC >= minimumTradeAmount):
                        print("Buy Stock")
                        print()
                        buyStock()
                    if(tC < minimumTradeAmount):
                        print("No Trade Due to Low Balance")
                        print()
                if((buySellList[stockIndex] == 'Buy') and (currentPriceFloat >= (averagePriceFloat * 1.05))):
                    print("Sell Stock")
                    print()
                    sellStock()
                if((buySellList[stockIndex] == 'Buy') and (currentPriceFloat <= (oldPriceFloat * .90))):
                    buyMore = True
                    if(tC >= minimumTradeAmount):
                        print("Buy More")
                        print()
                        buyMoreStock()
                    if(tC < minimumTradeAmount):
                        print("No Trade Due to Low Balance")
                        print()
                buyMore = False
                time.sleep(1)

            # Clear Lists to free up memory 
            currentPriceList.clear()
            oldPriceList.clear()
            quantityOwnedList.clear()
            recordedList.clear()
            shouldBeRecordedList.clear()
            
            print("End of Run")
            time.sleep(60)
            getDateTime = datetime.datetime.now()
            timeNow = getDateTime.strftime("%X")
            print(f"The Time of Day is: {timeNow}")
            print(f"Currently Trading {len(stockList)} stocks.")
            print(f"{len(tradedList)} trades executed this run.")
            
            print()
            print()            
            
            stockList.clear() # Clears the list from the last run.
            tradedList.clear()
            print()
            time.sleep(1)
            print("Start Time Was:", starttime)
            endtime = 0
            endtime = getDateTime
            print("End Time is:", endtime)
            totaltime = 0
            totaltime = endtime-starttime
            print(f'Total Execution Time: {totaltime}')
            print()
            print()
#########################################################################

        # Something to do after hours
        else:
            getDateTime = datetime.datetime.now()
            today = getDateTime.strftime(" %b %d, %Y")
            getDay = getDateTime.strftime("%A")
            getHour = getDateTime.strftime("%H")
            timeNow = getDateTime.strftime("%X")
            getMinute = getDateTime.strftime("%M")
            if((getHour >= '13') and (getHour <= '00')):
                print()
                print(f"Today is: {getDay} {today}")
                print(f"The Time of Day is: {timeNow}")
                print("After Hours -- Trading Closed")
                print()
                if(getMinute == '00'):
                    print()
                    print(f"Today is: {getDay} {today}")
                    print(f"The Time of Day is: {timeNow}")
                    print("After Hours -- Trading Closed")
                    print()
                    print()
                    print()
            time.sleep(1)
            
#########################################################################
    
    # Something to do on the weekends                
    else:
        getDateTime = datetime.datetime.now()
        today = getDateTime.strftime(" %b %d, %Y")
        getDay = getDateTime.strftime("%A")
        getMinute = getDateTime.strftime("%M")
        timeNow = getDateTime.strftime("%X")
        if(getMinute == '00'):
            print()
            print("Weekend -- Trading Closed")
            print(f"Today is: {getDay} {today}")
            print(f"The Time of Day is: {timeNow}")
            print()
            print()
            print()
            time.sleep(60)
        time.sleep(1)






